import os
import requests
import json
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import openpyxl
from openpyxl.styles import Alignment
from copy import copy
from datetime import datetime
from flask import Flask, request, abort, send_file
from linebot import LineBotApi, WebhookHandler
from linebot.exceptions import InvalidSignatureError
from linebot.models import MessageEvent, TextMessage, TextSendMessage

app = Flask(__name__)

# 🌟 ดึงกุญแจจากตู้เซฟ Secrets
LINE_CHANNEL_ACCESS_TOKEN = os.environ.get("LINE_ACCESS_TOKEN")
LINE_CHANNEL_SECRET = os.environ.get("LINE_SECRET")

line_bot_api = LineBotApi(LINE_CHANNEL_ACCESS_TOKEN)
handler = WebhookHandler(LINE_CHANNEL_SECRET)

# 🌟 ลิงก์ระบบเดิม (Google Apps Script)
GAS_WEBHOOK_URL = "https://script.google.com/macros/s/AKfycbwkGM0ryX5gCzjMvvDIKN3yS-gEHyNctoVB04kp8s8yJgPeMalwWOZJxfdS2fLTu18/exec"


def generate_excel_report(target_month, target_year):
    scope = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = ServiceAccountCredentials.from_json_keyfile_name("/etc/secrets/credentials.json", scope)
    client = gspread.authorize(creds)

    sheet_name = "Master Database ระบบจัดการน้ำมัน (LINE Bot)"
    sheet = client.open(sheet_name).sheet1
    all_data = sheet.get_all_values()

    fuel_data = {
        "ดีเซล": {"records": [], "open_bal": 0.0, "tot_in": 0.0, "tot_out": 0.0},
        "เบนซิน": {"records": [], "open_bal": 0.0, "tot_in": 0.0, "tot_out": 0.0},
    }

    for row in all_data[1:]:
        if len(row) < 10:
            continue

        date_str = str(row[1]).split(" ")[0].strip()
        if not date_str:
            continue
        try:
            if "/" in date_str:
                parts = date_str.split("/")
                date_val = datetime(int(parts[2]), int(parts[1]), int(parts[0]))
            elif "-" in date_str:
                parts = date_str.split("-")
                date_val = datetime(int(parts[0]), int(parts[1]), int(parts[2]))
            else:
                continue
        except Exception:
            continue

        row_year = date_val.year
        if row_year > 2500:
            row_year -= 543
        row_month = date_val.month
        type_val = str(row[2]).strip()

        target_fuel = None
        if "ดีเซล" in type_val:
            target_fuel = "ดีเซล"
        elif "เบนซิน" in type_val or "แก๊สโซฮอล์" in type_val:
            target_fuel = "เบนซิน"
        if not target_fuel:
            continue

        license_plate = str(row[3]).strip() if str(row[3]).strip() else "-"
        location = str(row[4]).strip() if str(row[4]).strip() else "-"
        mission = str(row[5]).strip() if str(row[5]).strip() else "-"

        in_str = str(row[6]).replace(",", "").strip()
        out_str = str(row[7]).replace(",", "").strip()
        in_amount = float(in_str) if in_str else 0.0
        out_amount = float(out_str) if out_str else 0.0

        price_str = str(row[8]).replace(",", "").strip()
        price = float(price_str) if price_str else "-"
        note = str(row[9]).strip() if len(row) > 9 and str(row[9]).strip() else ""

        if row_year < target_year or (
            row_year == target_year and row_month < target_month
        ):
            fuel_data[target_fuel]["open_bal"] += in_amount - out_amount
        elif row_year == target_year and row_month == target_month:
            fuel_data[target_fuel]["tot_in"] += in_amount
            fuel_data[target_fuel]["tot_out"] += out_amount
            thai_year = target_year + 543
            thai_yy_2digits = str(thai_year)[-2:]
            formatted_date = f"{date_val.day}/{date_val.month}/{thai_yy_2digits}"
            fuel_data[target_fuel]["records"].append(
                {
                    "date": formatted_date,
                    "license": license_plate,
                    "location": location,
                    "mission": mission,
                    "price": price,
                    "in_amount": in_amount,
                    "out_amount": out_amount,
                    "note": note,
                }
            )

    wb = openpyxl.load_workbook("template_oil.xlsx")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    def fill_sheet(ws, data, month, year):
        thai_months = [
            "",
            "มกราคม",
            "กุมภาพันธ์",
            "มีนาคม",
            "เมษายน",
            "พฤษภาคม",
            "มิถุนายน",
            "กรกฎาคม",
            "สิงหาคม",
            "กันยายน",
            "ตุลาคม",
            "พฤศจิกายน",
            "ธันวาคม",
        ]
        try:
            ws["B4"].value = f"ประจำเดือน {thai_months[month]} {year + 543}"
        except AttributeError:
            pass

        start_row = 7
        current_balance = round(data["open_bal"], 2)
        ws.cell(
            row=start_row, column=3, value="ยอดยกมาจากเดือนก่อน"
        ).alignment = right_align
        c_ob_val = ws.cell(row=start_row, column=10, value=current_balance)
        c_ob_val.number_format = "#,##0.00"
        c_ob_val.alignment = right_align

        current_row = start_row + 1
        idx = 1
        template_styles = {}
        for c in range(1, 11):
            t_cell = ws.cell(row=8, column=c)
            template_styles[c] = {
                "font": copy(t_cell.font),
                "border": copy(t_cell.border),
            }

        for r in data["records"]:
            in_amt = r["in_amount"]
            out_amt = r["out_amount"]
            current_balance = current_balance + in_amt - out_amt
            final_note = f"{round(current_balance, 2):,.2f}"
            if r["note"] and r["note"] != "-":
                final_note += f" ({r['note']})"

            cells_to_write = [
                (1, idx),
                (2, r["date"]),
                (3, r["license"]),
                (4, r["location"]),
                (5, r["mission"]),
                (6, r["price"]),
                (7, "-"),
                (8, in_amt if in_amt > 0 else "-"),
                (9, out_amt if out_amt > 0 else "-"),
                (10, final_note),
            ]

            for col, val in cells_to_write:
                c = ws.cell(row=current_row, column=col, value=val)
                style = template_styles[col]
                c.font = copy(style["font"])
                c.border = copy(style["border"])
                if col in [3, 5]:
                    c.alignment = left_align
                elif col in [6, 8, 9, 10]:
                    c.alignment = right_align
                else:
                    c.alignment = center_align
                if col in [8, 9] and isinstance(val, float):
                    c.number_format = "#,##0.00"

            current_row += 1
            idx += 1

        ws.cell(
            row=current_row, column=3, value="รวมใช้ประจำเดือน"
        ).alignment = right_align
        c_tot_out = ws.cell(
            row=current_row,
            column=9,
            value=data["tot_out"] if data["tot_out"] > 0 else "-",
        )
        c_tot_out.alignment = right_align
        if data["tot_out"] > 0:
            c_tot_out.number_format = "#,##0.00"

        ws.cell(
            row=current_row + 1, column=3, value="รับเพิ่มประจำเดือน"
        ).alignment = right_align
        c_tot_in = ws.cell(
            row=current_row + 1,
            column=8,
            value=data["tot_in"] if data["tot_in"] > 0 else "-",
        )
        c_tot_in.alignment = right_align
        if data["tot_in"] > 0:
            c_tot_in.number_format = "#,##0.00"

        ws.cell(
            row=current_row + 2, column=3, value="ยอดยกไปเดือนหน้า"
        ).alignment = right_align
        c_bal = ws.cell(row=current_row + 2, column=10, value=current_balance)
        c_bal.alignment = right_align
        c_bal.number_format = "#,##0.00"

        for r_idx in range(current_row, current_row + 3):
            for c_idx in range(1, 11):
                c = ws.cell(row=r_idx, column=c_idx)
                if c_idx in template_styles:
                    style = template_styles[c_idx]
                    if not c.font:
                        c.font = copy(style["font"])
                    if not c.border:
                        c.border = copy(style["border"])

    if "ดีเซล" in wb.sheetnames:
        fill_sheet(wb["ดีเซล"], fuel_data["ดีเซล"], target_month, target_year)
    if "เบนซิน" in wb.sheetnames:
        fill_sheet(wb["เบนซิน"], fuel_data["เบนซิน"], target_month, target_year)

    output_filename = f"Oil_Report_{target_month}_{target_year}.xlsx"
    wb.save(output_filename)
    return output_filename


@app.route("/webhook", methods=["POST"])
def webhook():
    signature = request.headers.get("X-Line-Signature", "")
    body = request.get_data(as_text=True)

    try:
        data = json.loads(body)
        events = data.get("events", [])

        is_report_command = False

        # เช็คว่ามีคนพิมพ์คำว่า "รายงาน" หรือไม่
        for event in events:
            if (
                event.get("type") == "message"
                and event.get("message", {}).get("type") == "text"
            ):
                if event["message"]["text"].strip() == "รายงาน":
                    is_report_command = True
                    break

        if is_report_command:
            # สั่งให้บอทจัดการสร้างรายงาน
            handler.handle(body, signature)
        else:
            # โยนข้อมูลไปให้ Google Apps Script
            headers = {
                "Content-Type": "application/json",
                "X-Line-Signature": signature,
            }
            requests.post(GAS_WEBHOOK_URL, data=body.encode("utf-8"), headers=headers)

    except Exception as e:
        print("Error in webhook processing:", e)

    return "OK"


# 🌟 ฟังก์ชันตอบกลับเวลาสั่ง "รายงาน"
@handler.add(MessageEvent, message=TextMessage)
def handle_message(event):
    text = event.message.text.strip()

    if text == "รายงาน":
        try:
            now = datetime.now()
            report_month = now.month - 1
            report_year = now.year
            if report_month == 0:
                report_month = 12
                report_year -= 1

            # สร้างไฟล์ Excel
            filename = generate_excel_report(report_month, report_year)

            # 🌟 แก้ไขตรงนี้ครับ (บรรทัดที่ 304-305 ในโค้ดเดิม)
            PUBLIC_URL = "https://oil-bot-7pgf.onrender.com"
            download_link = f"{PUBLIC_URL}/download/{filename}"

            line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(
                    text=f"✅ สร้างรายงานบัญชีน้ำมัน เดือน {report_month}/{report_year} สำเร็จแล้วครับ!\n\n📥 กดโหลดไฟล์ Excel ที่ลิงก์นี้ได้เลยครับ:\n{download_link}"
                ),
            )
        except Exception as e:
            # ถ้าเกิด Error จะเด้งบอกใน LINE ทันที
            error_msg = f"❌ ขออภัยครับ ระบบเกิดข้อผิดพลาดในการสร้างรายงาน:\n{str(e)}"
            line_bot_api.reply_message(
                event.reply_token, TextSendMessage(text=error_msg)
            )
            print(error_msg)


@app.route("/download/<filename>")
def download_file(filename):
    return send_file(filename, as_attachment=True)


@app.route("/")
def home():
    return "🚀 LINE Bot is Running 24/7!"


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8080)
